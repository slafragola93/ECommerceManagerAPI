"""Export CSV/Excel ricevute (BE-2.5)."""
from __future__ import annotations

import csv
import io
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from src.schemas.ricevuta_schema import (
    RicevutaAddressEmbedSchema,
    RicevutaListItemSchema,
    RicevutaResponseSchema,
)


class RicevutaExportService:
    DETAIL_HEADERS = [
        "numero",
        "anno",
        "data_emissione",
        "data_incasso",
        "stato",
        "id_ricevuta",
        "id_order",
        "order_reference",
        "customer_name",
        "customer_email",
        "address_country",
        "address_city",
        "product_reference",
        "product_name",
        "product_qty",
        "unit_price_net",
        "unit_price_with_tax",
        "total_price_net",
        "total_price_with_tax",
        "is_shipping",
        "order_total_net",
        "order_total_with_tax",
    ]

    LIST_HEADERS = [
        "id_ricevuta",
        "numero",
        "anno",
        "data_emissione",
        "data_incasso",
        "stato",
        "id_order",
        "order_reference",
        "customer_name",
        "customer_email",
        "order_total_with_tax",
    ]

    @staticmethod
    def _customer_name(
        firstname: Optional[str], lastname: Optional[str]
    ) -> str:
        return " ".join(part for part in (firstname, lastname) if part).strip()

    @classmethod
    def _primary_address(
        cls, ricevuta: RicevutaResponseSchema
    ) -> Optional[RicevutaAddressEmbedSchema]:
        return ricevuta.address_invoice or ricevuta.address_delivery

    @classmethod
    def _detail_rows(cls, ricevuta: RicevutaResponseSchema) -> List[list]:
        order = ricevuta.order
        customer = ricevuta.customer
        address = cls._primary_address(ricevuta)
        customer_name = cls._customer_name(
            customer.firstname if customer else None,
            customer.lastname if customer else None,
        )
        country = address.country.iso_code if address and address.country else None
        city = address.city if address else None

        base = [
            ricevuta.numero,
            ricevuta.anno,
            ricevuta.data_emissione.isoformat(),
            ricevuta.data_incasso.isoformat(),
            ricevuta.stato.value,
            ricevuta.id_ricevuta,
            order.id_order if order else None,
            order.reference if order else None,
            customer_name or None,
            customer.email if customer else None,
            country,
            city,
        ]
        order_total_net = order.total_price_net if order else None
        order_total_with_tax = order.total_price_with_tax if order else None

        if not ricevuta.order_details:
            return [
                base
                + [None, None, None, None, None, None, None, False, order_total_net, order_total_with_tax]
            ]

        rows: List[list] = []
        for line in ricevuta.order_details:
            rows.append(
                base
                + [
                    line.product_reference,
                    line.product_name,
                    line.product_qty,
                    line.unit_price_net,
                    line.unit_price_with_tax,
                    line.total_price_net,
                    line.total_price_with_tax,
                    line.is_shipping,
                    order_total_net,
                    order_total_with_tax,
                ]
            )
        return rows

    @classmethod
    def _list_row(cls, item: RicevutaListItemSchema) -> list:
        customer = item.customer
        return [
            item.id_ricevuta,
            item.numero,
            item.anno,
            item.data_emissione.isoformat(),
            item.data_incasso.isoformat(),
            item.stato.value,
            item.id_order,
            item.order_reference,
            cls._customer_name(
                customer.firstname if customer else None,
                customer.lastname if customer else None,
            )
            or None,
            customer.email if customer else None,
            item.order_total_with_tax,
        ]

    def build_detail_csv(self, ricevuta: RicevutaResponseSchema) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
        writer.writerow(self.DETAIL_HEADERS)
        for row in self._detail_rows(ricevuta):
            writer.writerow(row)
        return buffer.getvalue().encode("utf-8-sig")

    def build_list_csv(self, items: Iterable[RicevutaListItemSchema]) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
        writer.writerow(self.LIST_HEADERS)
        for item in items:
            writer.writerow(self._list_row(item))
        return buffer.getvalue().encode("utf-8-sig")

    def build_detail_xlsx(self, ricevuta: RicevutaResponseSchema) -> bytes:
        workbook = Workbook()
        header_sheet = workbook.active
        header_sheet.title = "Ricevuta"

        meta_rows = [
            ("Numero", f"{ricevuta.numero}/{ricevuta.anno}"),
            ("Data emissione", ricevuta.data_emissione.isoformat()),
            ("Data incasso", ricevuta.data_incasso.isoformat()),
            ("Stato", ricevuta.stato.value),
            ("ID ordine", ricevuta.order.id_order if ricevuta.order else None),
            (
                "Totale ordine lordo",
                ricevuta.order.total_price_with_tax if ricevuta.order else None,
            ),
        ]
        for label, value in meta_rows:
            header_sheet.append([label, value])

        lines_sheet = workbook.create_sheet("order_details")
        lines_sheet.append(self.DETAIL_HEADERS)
        for cell in lines_sheet[1]:
            cell.font = Font(bold=True)
        for row in self._detail_rows(ricevuta):
            lines_sheet.append(row)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def build_list_xlsx(self, items: Iterable[RicevutaListItemSchema]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ricevute"
        sheet.append(self.LIST_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for item in items:
            sheet.append(self._list_row(item))

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
