"""Test export Excel corrispettivi — matrice aliquote con IVA."""
from datetime import date
from decimal import Decimal
from io import BytesIO

import zipfile
from openpyxl import load_workbook

from src.schemas.corrispettivo_schema import (
    CorrispettivoRiepilogoResponseSchema,
    CorrispettivoRiepilogoRowSchema,
    CorrispettivoRiepilogoTotalsSchema,
    CorrispettivoTaxCellSchema,
    CorrispettivoTaxColumnSchema,
)
from src.services.export.corrispettivi_excel_service import CorrispettiviExcelService


def _sample_riepilogo(*, delivery_country_iso: str | None = None) -> CorrispettivoRiepilogoResponseSchema:
    return CorrispettivoRiepilogoResponseSchema(
        year=2026,
        month=7,
        delivery_country_iso=delivery_country_iso,
        columns=[
            CorrispettivoTaxColumnSchema(id_tax=1, label="22", percentage=22.0),
            CorrispettivoTaxColumnSchema(id_tax=9, label="0", percentage=0.0),
        ],
        rows=[
            CorrispettivoRiepilogoRowSchema(
                day=8,
                date=date(2026, 7, 8),
                cells={
                    "1": CorrispettivoTaxCellSchema(
                        products_sales=Decimal("100.00"),
                        shipping_sales=Decimal("12.00"),
                        products_returns=Decimal("10.00"),
                        shipping_returns=Decimal("2.00"),
                    ),
                    "9": CorrispettivoTaxCellSchema(products_sales=Decimal("20.00")),
                },
                row_total=Decimal("120.00"),
            )
        ],
        month_totals=CorrispettivoRiepilogoTotalsSchema(
            products_sales=Decimal("120.00"),
            shipping_sales=Decimal("12.00"),
            products_returns=Decimal("10.00"),
            shipping_returns=Decimal("2.00"),
            row_total=Decimal("120.00"),
        ),
    )


class TestCorrispettiviExcelService:
    def test_riepilogo_workbook_single_day_footer(self):
        riepilogo = _sample_riepilogo()
        riepilogo.day = 8
        raw = CorrispettiviExcelService().build_riepilogo_workbook(riepilogo)
        sheet = load_workbook(BytesIO(raw)).active

        assert sheet.cell(3, 1).value == "Totale 08/07/2026"

    def test_riepilogo_workbook_includes_tax_columns(self):
        raw = CorrispettiviExcelService().build_riepilogo_workbook(_sample_riepilogo())
        sheet = load_workbook(BytesIO(raw)).active

        assert sheet.title == "Riepilogo"
        assert sheet.cell(1, 1).value == "Data"
        assert sheet.cell(1, 2).value == "22 - Totale entrate prodotti"
        assert sheet.cell(1, 6).value == "0 - Totale entrate prodotti"
        assert sheet.cell(1, 9).value == "0 - Totale resi spedizione"
        assert sheet.cell(2, 1).value == "08/07/2026"
        assert sheet.cell(2, 2).value == 100.0
        assert sheet.cell(3, 1).value == "Totale 07/2026"
        assert sheet.cell(3, 5).value == 2.0

    def test_zip_contains_registro_files(self):
        consolidated = _sample_riepilogo()
        it_riepilogo = _sample_riepilogo(delivery_country_iso="IT")

        raw = CorrispettiviExcelService().build_registri_zip(
            consolidated_riepilogo=consolidated,
            by_country={"IT": it_riepilogo},
        )

        with zipfile.ZipFile(BytesIO(raw)) as archive:
            names = set(archive.namelist())
            it_sheet = load_workbook(
                BytesIO(archive.read("registro_IT.xlsx"))
            ).active

        assert "registro.xlsx" in names
        assert "registro_IT.xlsx" in names
        assert it_sheet.cell(1, 2).value == "22 - Totale entrate prodotti"
        assert it_sheet.cell(1, 6).value == "0 - Totale entrate prodotti"

    def test_country_workbook_same_structure_as_consolidated(self):
        consolidated = _sample_riepilogo()
        fr_riepilogo = _sample_riepilogo(delivery_country_iso="FR")

        service = CorrispettiviExcelService()
        consolidated_sheet = load_workbook(
            BytesIO(service.build_riepilogo_workbook(consolidated))
        ).active
        country_sheet = load_workbook(
            BytesIO(service.build_riepilogo_workbook(fr_riepilogo))
        ).active

        consolidated_headers = [cell.value for cell in consolidated_sheet[1]]
        country_headers = [cell.value for cell in country_sheet[1]]
        assert consolidated_headers == country_headers
