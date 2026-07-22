"""Test export massivo fatture."""
from __future__ import annotations

import io
import zipfile
from datetime import datetime

import pytest
from openpyxl import load_workbook

from src.schemas.fiscal_document_schema import (
    InvoiceExportFiltersSchema,
    InvoiceListExportItemSchema,
)
from src.services.export.fiscal_document_export_service import FiscalDocumentExportService


def _sample_item(**overrides) -> InvoiceListExportItemSchema:
    data = {
        "id_fiscal_document": 1,
        "document_type": "invoice",
        "document_number": "000001",
        "internal_number": None,
        "tipo_documento_fe": "TD01",
        "status": "issued",
        "is_electronic": True,
        "id_order": 100,
        "order_reference": "ORD-100",
        "customer_firstname": "Mario",
        "customer_lastname": "Rossi",
        "customer_email": "mario@example.com",
        "delivery_country_iso": "IT",
        "delivery_city": "Roma",
        "date_add": datetime(2026, 1, 15, 10, 30, 0),
        "total_price_net": 100.0,
        "total_price_with_tax": 122.0,
        "products_total_price_net": 90.0,
        "products_total_price_with_tax": 109.8,
    }
    data.update(overrides)
    return InvoiceListExportItemSchema(**data)


class TestFiscalDocumentExportService:
    def setup_method(self):
        self.service = FiscalDocumentExportService()

    def test_build_list_xlsx_headers_and_row(self):
        content = self.service.build_list_xlsx([_sample_item()])
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active

        assert sheet.title == "Fatture"
        assert sheet.max_row == 2
        assert sheet.cell(row=1, column=1).value == "id_fiscal_document"
        assert sheet.cell(row=1, column=2).value == "document_type"
        assert sheet.cell(row=2, column=2).value == "invoice"
        assert sheet.cell(row=2, column=3).value == "000001"
        assert sheet.cell(row=2, column=10).value == "Mario Rossi"
        assert sheet.cell(row=2, column=15).value == 100.0

    def test_build_xml_zip_contains_files(self):
        def fake_loader(invoice_id: int) -> tuple[bytes, str]:
            return b"<?xml version='1.0'?>", f"IT01234567890_{invoice_id:05d}.xml"

        content = self.service.build_xml_zip([1, 2], fake_loader)
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = sorted(archive.namelist())
            assert names == ["IT01234567890_00001.xml", "IT01234567890_00002.xml"]


class TestInvoiceExportFiltersSchema:
    def test_invalid_date_range(self):
        with pytest.raises(ValueError, match="date_add_to"):
            InvoiceExportFiltersSchema(
                date_add_from=datetime(2026, 2, 1).date(),
                date_add_to=datetime(2026, 1, 1).date(),
            )

    def test_normalizes_delivery_country_iso(self):
        filters = InvoiceExportFiltersSchema(delivery_country_iso="it")
        assert filters.delivery_country_iso == "IT"

    def test_normalizes_document_type(self):
        filters = InvoiceExportFiltersSchema(document_type="CREDIT_NOTE")
        assert filters.document_type == "credit_note"

    def test_for_xml_export_strips_extra_filters(self):
        source = InvoiceExportFiltersSchema(
            is_electronic=True,
            status="pending",
            id_order=10,
            id_customer=20,
            delivery_country_iso="DE",
            date_add_from=datetime(2026, 1, 1).date(),
            date_add_to=datetime(2026, 1, 31).date(),
        )
        xml_filters = source.for_xml_export(max_limit=5000)
        assert xml_filters.is_electronic is True
        assert xml_filters.status is None
        assert xml_filters.id_order is None
        assert xml_filters.id_customer is None
        assert xml_filters.delivery_country_iso == "DE"
        assert xml_filters.date_add_from.isoformat() == "2026-01-01"
        assert xml_filters.limit == 5000
